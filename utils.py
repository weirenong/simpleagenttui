from __future__ import annotations
from math import sqrt
import hashlib
from pathlib import Path
import difflib
import editblock
import base64
import urllib.error
import urllib.request
import html
from html.parser import HTMLParser
import urllib.parse
import re
import csv
import json
from pathlib import Path
import zipfile
import xml.etree.ElementTree as ET
from typing import Any, Callable, Sequence, TypeVar, cast, overload
import mimetypes
from urllib.parse import quote_plus, urlparse, parse_qs, unquote
from playwright.sync_api import Browser, BrowserContext, Page, sync_playwright
try:
    from langchain_text_splitters import (
        Language,
        MarkdownHeaderTextSplitter,
        RecursiveCharacterTextSplitter,
        RecursiveJsonSplitter,
    )
except ImportError:  # pragma: no cover
    Language = None
    MarkdownHeaderTextSplitter = None
    RecursiveCharacterTextSplitter = None
    RecursiveJsonSplitter = None
try:
    import faiss
except ImportError:  # pragma: no cover
    faiss = None
try:
    import numpy as np
except ImportError:  # pragma: no cover
    np = None

try:
    from PIL import Image
except ImportError:  # pragma: no cover
    Image = None

try:
    import pypdf
except ImportError:  # pragma: no cover
    pypdf = None

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

T = TypeVar("T", bound=dict)


def calculate_file_sha256(path: str | Path) -> str:
    """
    Calculate SHA-256 hash for a file.
    """
    file_path = Path(path).expanduser()
    sha256 = hashlib.sha256()
    with file_path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            sha256.update(chunk)
    return sha256.hexdigest()


# -----------------------------
# Hidden model prompt utilities
# -----------------------------
def send_status_message(
    message: str | None,
    status_printer: Callable[[str], None] | None = None,
) -> None:
    """
    Send a user-facing status message through the caller's UI printer.
    """
    message = (message or "").strip()
    if not message or status_printer is None:
        return

    status_printer(message)

def prompt_text_model(
    client: Any,
    prompt: str,
    model: str,
    system_prompt: str | None = None,
    temperature: float | None = None,
    top_p: float | None = None,
    options: dict[str, Any] | None = None,
) -> str:
    """
    Run a hidden/internal prompt through the main text model and return plain text.
    """
    prompt = (prompt or "").strip()
    if not prompt:
        return ""

    messages: list[dict[str, str]] = []
    if system_prompt:
        messages.append({"role": "system", "content": system_prompt.strip()})

    messages.append({"role": "user", "content": prompt})

    generation_options = dict(options or {})
    if temperature is not None:
        generation_options["temperature"] = temperature
    if top_p is not None:
        generation_options["top_p"] = top_p

    chat_kwargs: dict[str, Any] = {
        "stream": True,
        "model": model,
    }

    if generation_options:
        chat_kwargs["options"] = generation_options

    response_text = ""
    try:
        response_stream = client.chat(messages, **chat_kwargs)
    except TypeError:
        # Fallback for lightweight wrappers that do not accept Ollama options yet.
        response_stream = client.chat(messages, stream=True, model=model)

    for chunk in response_stream:
        response_text += normalise_model_text_response(chunk)

    return response_text.strip()


def prompt_vision_model(
    client: Any,
    prompt: str,
    image_paths: Sequence[str],
    model: str,
    system_prompt: str | None = None,
    temperature: float | None = None,
    top_p: float | None = None,
    options: dict[str, Any] | None = None,
) -> str:
    """
    Run a hidden/internal prompt through a vision model and return plain text.

    image_paths should contain local image file paths. The client is expected to
    support Ollama-style multimodal chat messages where the user message includes
    an `images` list. This function does not print anything to the TUI.
    """
    prompt = (prompt or "").strip()
    valid_image_paths = [str(path).strip() for path in image_paths if str(path).strip()]

    if not prompt or not valid_image_paths:
        return ""

    messages: list[dict[str, Any]] = []
    if system_prompt:
        messages.append({"role": "system", "content": system_prompt.strip()})

    messages.append(
        {
            "role": "user",
            "content": prompt,
            "images": valid_image_paths,
        }
    )

    generation_options = dict(options or {})
    if temperature is not None:
        generation_options["temperature"] = temperature
    if top_p is not None:
        generation_options["top_p"] = top_p

    chat_kwargs: dict[str, Any] = {
        "stream": True,
        "model": model,
    }

    if generation_options:
        chat_kwargs["options"] = generation_options

    response_text = ""
    try:
        response_stream = client.chat(messages, **chat_kwargs)
    except TypeError:
        response_stream = client.chat(messages, stream=True, model=model)

    for chunk in response_stream:
        response_text += normalise_model_text_response(chunk)

    return response_text.strip()


def normalise_model_text_response(response: Any) -> str:
    """
    Convert common Ollama/client response shapes into plain text.

    Supported shapes:
    - plain string chunks
    - {"message": {"content": "..."}}
    - {"response": "..."}
    - {"content": "..."}
    - objects with `.message.content`, `.response`, or `.content`
    """
    if response is None:
        return ""

    if isinstance(response, str):
        return response

    if isinstance(response, dict):
        message = response.get("message")
        if isinstance(message, dict):
            return str(message.get("content") or "")
        if response.get("response") is not None:
            return str(response.get("response") or "")
        if response.get("content") is not None:
            return str(response.get("content") or "")
        return ""

    message = getattr(response, "message", None)
    if message is not None:
        content = getattr(message, "content", None)
        if content is not None:
            return str(content)

    for attribute in ("response", "content"):
        value = getattr(response, attribute, None)
        if value is not None:
            return str(value)

    return str(response)

def normalise_llm_file_headers(llm_output: str) -> str:
    """
    Clean common markdown/prose file labels before whole-file edit parsing.

    This keeps normal code blocks unchanged, but rewrites noisy labels such as
    `File name:** `hello.py`` into `File: hello.py` so downstream diff review
    can extract the intended path instead of the surrounding markdown.
    """
    if not llm_output:
        return ""

    lines = llm_output.splitlines()
    normalised_lines: list[str] = []

    for line in lines:
        filename = extract_filename_from_llm_header(line)
        if filename:
            normalised_lines.append(f"File: {filename}")
            continue

        cleaned_line = line.strip().strip("*`")
        if re.match(r"^\s*(file\s+contents?|contents?)\s*:\s*$", cleaned_line, flags=re.IGNORECASE):
            continue

        normalised_lines.append(line)

    return "\n".join(normalised_lines).strip()


def extract_filename_from_llm_header(line: str) -> str:
    """
    Extract a filename from simple LLM file headers.

    Keep this intentionally small: remove markdown decoration, recognise labels
    like `File name:` / `Filename:` / `File:`, then return the first plausible
    path-looking token after the colon.
    """
    value = line.strip()
    if not value:
        return ""

    value = value.strip("*` ")
    value = re.sub(r"\*+", "", value)

    match = re.match(
        r"^file\s*(?:name|path)?\s*:\s*(?P<rest>.+)$",
        value,
        flags=re.IGNORECASE,
    )
    if not match:
        return ""

    rest = match.group("rest").strip().strip("`'\"* ")
    if not rest:
        return ""

    token_match = re.search(r"[^\s`'\"<>|]+\.[A-Za-z0-9_.-]+", rest)
    if not token_match:
        return ""

    return token_match.group(0).strip("`'\"* ,:;")

# -----------------------------
# Embedding utilities
# -----------------------------

@overload
def vectorise_text(client: Any, data: str, model: str) -> list[float]:
    ...

@overload
def vectorise_text(client: Any, data: Sequence[str], model: str) -> list[list[float]]:
    ...

def vectorise_text(client: Any, data: str | Sequence[str], model: str) -> list[float] | list[list[float]]:
    """
    Convert one text item, or many text items, into embedding vectors.

    The client is expected to expose:
        client.embed(text, model=model)
    """
    if isinstance(data, str):
        return _vectorise_one_text(client, data, model)

    return [_vectorise_one_text(client, text, model) for text in data]

def _vectorise_one_text(client: Any, text: str, model: str) -> list[float]:
    text = (text or "").strip()
    if not text:
        return []

    embedding = client.embed(text, model=model)
    return normalise_embedding_vector(embedding)


def normalise_embedding_vector(embedding: Any) -> list[float]:
    """
    Normalise common embedding response shapes into a plain list[float].

    Supported shapes:
    - [0.1, 0.2, ...]
    - {"embedding": [...]}
    - {"embeddings": [[...]]}
    """
    if isinstance(embedding, dict):
        if "embedding" in embedding:
            embedding = embedding["embedding"]
        elif "embeddings" in embedding and embedding["embeddings"]:
            embedding = embedding["embeddings"][0]

    if not isinstance(embedding, Sequence) or isinstance(embedding, (str, bytes)):
        return []

    vector: list[float] = []
    for value in embedding:
        try:
            vector.append(float(value))
        except (TypeError, ValueError):
            return []

    return vector

def is_embedding_vector(value: object) -> bool:
    """
    Return True when value is one flat embedding vector.
    """
    if not isinstance(value, list) or not value:
        return False

    return all(isinstance(item, (int, float)) for item in value)

def normalise_vector_for_similarity(vector: Sequence[float]) -> list[float]:
    """
    Return an L2-normalised vector.

    FAISS IndexFlatIP + normalised vectors gives cosine-similarity ranking.
    """
    if not vector:
        return []

    square_sum = sum(value * value for value in vector)
    if square_sum <= 0:
        return []

    norm = sqrt(square_sum)
    return [float(value) / norm for value in vector]


def cosine_similarity(left: Sequence[float], right: Sequence[float]) -> float:
    """
    Fallback comparison for already-embedded vectors.
    """
    normalised_left = normalise_vector_for_similarity(left)
    normalised_right = normalise_vector_for_similarity(right)

    if not normalised_left or not normalised_right or len(normalised_left) != len(normalised_right):
        return 0.0

    return sum(
        left_value * right_value
        for left_value, right_value in zip(normalised_left, normalised_right)
    )


class VectorMemoryIndex:
    """
    FAISS-backed in-memory vector index.

    This can index:
    - compacted conversation memory
    - webpage chunks
    - attachment chunks
    - code snippets
    """

    def __init__(self, embedding_key: str = "embedding") -> None:
        self.embedding_key = embedding_key
        self.items: list[dict[str, Any]] = []
        self.index = None
        self.dimension: int | None = None
        self.backend = "faiss" if faiss is not None and np is not None else "cosine"

    def clear(self) -> None:
        self.items.clear()
        self.index = None
        self.dimension = None

    def add_item(self, item: dict[str, Any]) -> bool:
        embedding = normalise_embedding_vector(item.get(self.embedding_key) or [])
        normalised_embedding = normalise_vector_for_similarity(embedding)

        if not normalised_embedding:
            return False

        item[self.embedding_key] = normalised_embedding

        if self.dimension is None:
            self.dimension = len(normalised_embedding)
            self._create_faiss_index()

        if len(normalised_embedding) != self.dimension:
            return False

        self.items.append(item)

        if self.backend == "faiss" and self.index is not None:
            vector = np.array([normalised_embedding], dtype="float32")
            self.index.add(vector)

        return True

    def rebuild(self, items: Sequence[dict[str, Any]]) -> None:
        self.clear()
        for item in items:
            self.add_item(dict(item))

    def search(
        self,
        query_embedding: Sequence[float],
        top_k: int = 8,
        min_score: float = 0.05,
    ) -> list[dict[str, Any]]:
        if not self.items or top_k <= 0:
            return []

        query_vector = normalise_vector_for_similarity(query_embedding)
        if not query_vector or self.dimension is None or len(query_vector) != self.dimension:
            return []

        if self.backend == "faiss" and self.index is not None:
            return self._search_faiss(query_vector, top_k=top_k, min_score=min_score)

        return self._search_cosine(query_vector, top_k=top_k, min_score=min_score)

    def _create_faiss_index(self) -> None:
        if self.backend != "faiss" or self.dimension is None:
            return

        # Inner product on L2-normalised vectors is equivalent to cosine similarity.
        self.index = faiss.IndexFlatIP(self.dimension)

    def _search_faiss(
        self,
        query_vector: Sequence[float],
        top_k: int,
        min_score: float,
    ) -> list[dict[str, Any]]:
        search_k = min(top_k, len(self.items))
        vector = np.array([query_vector], dtype="float32")
        scores, indices = self.index.search(vector, search_k)

        results: list[dict[str, Any]] = []
        for score, index in zip(scores[0], indices[0]):
            if index < 0:
                continue
            if float(score) < min_score:
                continue

            item = dict(self.items[int(index)])
            item["similarity_score"] = float(score)
            results.append(item)

        return results

    def _search_cosine(
        self,
        query_vector: Sequence[float],
        top_k: int,
        min_score: float,
    ) -> list[dict[str, Any]]:
        scored_items: list[tuple[float, int, dict[str, Any]]] = []

        for index, item in enumerate(self.items):
            context_embedding = item.get(self.embedding_key) or []
            score = cosine_similarity(query_vector, context_embedding)
            if score >= min_score:
                scored_items.append((score, index, item))

        scored_items.sort(
            key=lambda scored_item: (scored_item[0], -scored_item[1]),
            reverse=True,
        )

        results: list[dict[str, Any]] = []
        for score, _, item in scored_items[:top_k]:
            result = dict(item)
            result["similarity_score"] = score
            results.append(result)

        return results


def rank_embedded_contexts(
    query_embedding: Sequence[float],
    context_items: Sequence[T],
    embedding_key: str = "embedding",
    top_k: int = 8,
    min_score: float = 0.05,
) -> list[T]:
    """
    Compatibility helper.

    New code should prefer VectorMemoryIndex directly.
    """
    index = VectorMemoryIndex(embedding_key=embedding_key)
    index.rebuild(context_items)
    return index.search(query_embedding, top_k=top_k, min_score=min_score)


# -----------------------------
# Context-building helpers
# -----------------------------

# -----------------------------
# LangChain-backed chunking helpers
# -----------------------------

CODE_ATTACHMENT_EXTENSIONS = {
    ".py", ".js", ".jsx", ".ts", ".tsx", ".java", ".cs", ".cpp", ".c", ".h", ".hpp",
    ".go", ".rs", ".php", ".rb", ".swift", ".kt", ".kts", ".sql", ".sh", ".bash",
    ".zsh", ".ps1", ".bat", ".dockerfile",
}

MARKDOWN_ATTACHMENT_EXTENSIONS = {".md", ".markdown", ".rst"}

LANGCHAIN_LANGUAGE_BY_EXTENSION: dict[str, str] = {
    ".py": "PYTHON",
    ".js": "JS",
    ".jsx": "JS",
    ".ts": "TS",
    ".tsx": "TS",
    ".java": "JAVA",
    ".cs": "CSHARP",
    ".cpp": "CPP",
    ".c": "C",
    ".h": "C",
    ".hpp": "CPP",
    ".go": "GO",
    ".rs": "RUST",
    ".php": "PHP",
    ".rb": "RUBY",
    ".swift": "SWIFT",
    ".kt": "KOTLIN",
    ".kts": "KOTLIN",
    ".sql": "SQL",
    ".html": "HTML",
    ".htm": "HTML",
    ".md": "MARKDOWN",
    ".markdown": "MARKDOWN",
}


def estimate_token_count(text: str) -> int:
    """
    Lightweight model-agnostic token estimate for stats and metadata.
    """
    if not text:
        return 0

    return len(
        re.findall(
            r"[\u4e00-\u9fff]|[A-Za-z]+(?:'[A-Za-z]+)?|\d+(?:\.\d+)?|[^\s\w]",
            text,
            flags=re.UNICODE,
        )
    )


def chunk_text(
    text: str,
    chunk_size: int = 1_200,
    chunk_overlap: int = 180,
) -> list[str]:
    """
    Split generic text using LangChain's recursive splitter.

    Falls back to a small internal splitter only when langchain-text-splitters is not installed.
    """
    chunks = split_with_recursive_text_splitter(
        text=text,
        chunk_size=chunk_size,
        chunk_overlap=chunk_overlap,
    )
    return [chunk.strip() for chunk in chunks if chunk.strip()]


def chunk_source_text(
    text: str,
    source_path: str,
    extension: str,
    chunk_size: int = 1_200,
    chunk_overlap: int = 180,
) -> list[dict[str, Any]]:
    """
    Chunk source text with LangChain splitters and return rich chunk records.
    """
    extension = extension.lower()

    if extension in {".json", ".jsonl"}:
        return chunk_json_with_langchain(
            text=text,
            source_path=source_path,
            extension=extension,
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
        )

    if extension in MARKDOWN_ATTACHMENT_EXTENSIONS:
        return chunk_markdown_with_langchain(
            text=text,
            source_path=source_path,
            extension=extension,
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
        )

    if extension in CODE_ATTACHMENT_EXTENSIONS or extension in LANGCHAIN_LANGUAGE_BY_EXTENSION:
        return chunk_code_with_langchain(
            text=text,
            source_path=source_path,
            extension=extension,
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
        )

    chunks = split_with_recursive_text_splitter(
        text=text,
        chunk_size=chunk_size,
        chunk_overlap=chunk_overlap,
    )
    return build_chunk_records(
        chunks=chunks,
        chunk_kind="text",
        source_path=source_path,
        extension=extension,
    )


def split_with_recursive_text_splitter(
    text: str,
    chunk_size: int = 1_200,
    chunk_overlap: int = 180,
    separators: list[str] | None = None,
) -> list[str]:
    if not text or not text.strip():
        return []

    if RecursiveCharacterTextSplitter is None:
        return fallback_recursive_split(text, chunk_size=chunk_size, chunk_overlap=chunk_overlap)

    splitter_kwargs: dict[str, Any] = {
        "chunk_size": chunk_size,
        "chunk_overlap": chunk_overlap,
    }

    if separators:
        splitter_kwargs["separators"] = separators

    splitter = RecursiveCharacterTextSplitter(**splitter_kwargs)
    return splitter.split_text(text)


def chunk_markdown_with_langchain(
    text: str,
    source_path: str,
    extension: str,
    chunk_size: int = 1_200,
    chunk_overlap: int = 180,
) -> list[dict[str, Any]]:
    if MarkdownHeaderTextSplitter is None:
        chunks = split_with_recursive_text_splitter(text, chunk_size=chunk_size, chunk_overlap=chunk_overlap)
        return build_chunk_records(chunks, "markdown", source_path, extension)

    headers_to_split_on = [
        ("#", "Header 1"),
        ("##", "Header 2"),
        ("###", "Header 3"),
        ("####", "Header 4"),
    ]
    markdown_splitter = MarkdownHeaderTextSplitter(headers_to_split_on=headers_to_split_on)
    documents = markdown_splitter.split_text(text)
    records: list[dict[str, Any]] = []

    for document in documents:
        section_title = build_markdown_section_title(document.metadata)
        section_text = document.page_content

        if section_title:
            section_text = f"Section: {section_title}\n{section_text}"

        chunks = split_with_recursive_text_splitter(
            section_text,
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
        )
        records.extend(
            build_chunk_records(
                chunks=chunks,
                chunk_kind="markdown_section",
                source_path=source_path,
                extension=extension,
                extra_metadata={
                    "markdown_headers": dict(document.metadata),
                    "section_title": section_title,
                },
            )
        )

    return records


def build_markdown_section_title(metadata: dict[str, Any]) -> str:
    ordered_values = [
        str(value).strip()
        for _, value in sorted(metadata.items())
        if str(value).strip()
    ]
    return " > ".join(ordered_values)


def chunk_code_with_langchain(
    text: str,
    source_path: str,
    extension: str,
    chunk_size: int = 1_200,
    chunk_overlap: int = 180,
) -> list[dict[str, Any]]:
    splitter = create_language_splitter(extension, chunk_size=chunk_size, chunk_overlap=chunk_overlap)

    if splitter is None:
        chunks = split_with_recursive_text_splitter(
            text,
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
            separators=[
                "\nclass ",
                "\ndef ",
                "\nfunction ",
                "\nexport ",
                "\npublic ",
                "\nprivate ",
                "\n\n",
                "\n",
                " ",
                "",
            ],
        )
    else:
        chunks = splitter.split_text(text)

    return build_chunk_records(
        chunks=chunks,
        chunk_kind="code",
        source_path=source_path,
        extension=extension,
    )


def create_language_splitter(
    extension: str,
    chunk_size: int = 1_200,
    chunk_overlap: int = 180,
):
    if RecursiveCharacterTextSplitter is None or Language is None:
        return None

    language_name = LANGCHAIN_LANGUAGE_BY_EXTENSION.get(extension.lower())
    if not language_name:
        return None

    raw_language = getattr(Language, language_name, None)
    if raw_language is None:
        return None

    language = cast(Any, raw_language)

    try:
        return RecursiveCharacterTextSplitter.from_language(
            language=language,
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
        )
    except Exception:
        return None


def chunk_json_with_langchain(
    text: str,
    source_path: str,
    extension: str,
    chunk_size: int = 1_200,
    chunk_overlap: int = 180,
) -> list[dict[str, Any]]:
    if extension == ".jsonl":
        chunks = split_with_recursive_text_splitter(
            text,
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
            separators=["\n", " ", ""],
        )
        return build_chunk_records(chunks, "jsonl", source_path, extension)

    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        chunks = split_with_recursive_text_splitter(
            text,
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
        )
        return build_chunk_records(chunks, "json_fallback", source_path, extension)

    if RecursiveJsonSplitter is None:
        rendered = json.dumps(parsed, indent=2, ensure_ascii=False)
        chunks = split_with_recursive_text_splitter(
            rendered,
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
        )
        return build_chunk_records(chunks, "json", source_path, extension)

    splitter = RecursiveJsonSplitter(max_chunk_size=chunk_size)
    json_chunks = splitter.split_json(json_data=parsed)
    rendered_chunks = [
        json.dumps(chunk, indent=2, ensure_ascii=False)
        for chunk in json_chunks
    ]
    return build_chunk_records(rendered_chunks, "json", source_path, extension)


def build_chunk_records(
    chunks: Sequence[str],
    chunk_kind: str,
    source_path: str,
    extension: str,
    extra_metadata: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    metadata = extra_metadata or {}

    for index, chunk in enumerate(chunks):
        content = str(chunk or "").strip()
        if not content:
            continue

        records.append(
            {
                "content": content,
                "chunk_kind": chunk_kind,
                "splitter": "langchain-text-splitters" if RecursiveCharacterTextSplitter is not None else "fallback",
                "source_path": source_path,
                "extension": extension,
                "local_chunk_index": index,
                "estimated_tokens": estimate_token_count(content),
                **metadata,
            }
        )

    return records


def fallback_recursive_split(
    text: str,
    chunk_size: int = 1_200,
    chunk_overlap: int = 180,
) -> list[str]:
    text = (text or "").strip()
    if not text:
        return []

    if len(text) <= chunk_size:
        return [text]

    separators = ["\n\n", "\n", ". ", " ", ""]
    return fallback_split_by_separators(text, separators, chunk_size, chunk_overlap)


def fallback_split_by_separators(
    text: str,
    separators: Sequence[str],
    chunk_size: int,
    chunk_overlap: int,
) -> list[str]:
    if len(text) <= chunk_size:
        return [text.strip()]

    separator = separators[0]
    remaining_separators = separators[1:]

    if separator:
        parts = text.split(separator)
    else:
        parts = list(text)

    chunks: list[str] = []
    current = ""

    for part in parts:
        candidate = part if not current else f"{current}{separator}{part}"

        if len(candidate) <= chunk_size:
            current = candidate
            continue

        if current:
            chunks.append(current.strip())
            current = ""

        if len(part) > chunk_size and remaining_separators:
            chunks.extend(
                fallback_split_by_separators(
                    part,
                    remaining_separators,
                    chunk_size,
                    chunk_overlap,
                )
            )
        else:
            chunks.append(part[:chunk_size].strip())

    if current:
        chunks.append(current.strip())

    if chunk_overlap <= 0 or len(chunks) <= 1:
        return [chunk for chunk in chunks if chunk]

    overlapped_chunks: list[str] = []

    for index, chunk in enumerate(chunks):
        if index == 0:
            overlapped_chunks.append(chunk)
            continue

        previous_tail = chunks[index - 1][-chunk_overlap:]
        overlapped_chunks.append(f"{previous_tail}\n{chunk}".strip())

    return [chunk for chunk in overlapped_chunks if chunk]

def build_context_items_from_text(
    text: str,
    source_type: str,
    source_path: str,
    title: str | None = None,
    metadata: dict[str, Any] | None = None,
    chunk_size: int = 1_200,
    chunk_overlap: int = 180,
) -> list[dict[str, Any]]:
    """
    Convert raw source text into standard context items using LangChain-backed splitters.
    """
    base_metadata = metadata or {}
    extension = str(base_metadata.get("extension") or Path(source_path).suffix or "").lower()

    chunk_records = chunk_source_text(
        text=text,
        source_path=source_path,
        extension=extension,
        chunk_size=chunk_size,
        chunk_overlap=chunk_overlap,
    )

    context_items: list[dict[str, Any]] = []

    for index, record in enumerate(chunk_records):
        content = str(record.get("content") or "").strip()
        if not content:
            continue

        extra_metadata = {
            key: value
            for key, value in record.items()
            if key != "content"
        }

        context_items.append(
            {
                "source_type": source_type,
                "source_path": source_path,
                "title": title or source_path,
                "chunk_index": index,
                "content": content,
                "embedding": [],
                "estimated_tokens": estimate_token_count(content),
                **base_metadata,
                **extra_metadata,
            }
        )

    return context_items


def vectorise_context_items(
    client: Any,
    context_items: Sequence[dict[str, Any]],
    model: str,
    content_key: str = "content",
    embedding_key: str = "embedding",
) -> list[dict[str, Any]]:
    """
    Add embedding vectors to context items using their content field.

    This is the reusable bridge for conversation memory, webpages, attachments,
    and future code chunks. It returns copied dictionaries so callers do not
    accidentally mutate source data unless they explicitly store the result.
    """
    embedded_items: list[dict[str, Any]] = []

    for item in context_items:
        copied_item = dict(item)
        content = str(copied_item.get(content_key) or "").strip()

        if not content:
            copied_item[embedding_key] = []
            embedded_items.append(copied_item)
            continue

        embedding = vectorise_text(client, content, model)
        copied_item[embedding_key] = embedding if is_embedding_vector(embedding) else []
        embedded_items.append(copied_item)

    return embedded_items


# -----------------------------
# Attachment reading helpers
# -----------------------------

TEXT_ATTACHMENT_EXTENSIONS = {
    ".txt", ".md", ".markdown", ".rst", ".log", ".json", ".jsonl", ".yaml", ".yml", ".toml", ".ini", ".cfg", ".conf", ".env", ".gitignore",
    ".py", ".js", ".jsx", ".ts", ".tsx", ".html", ".css", ".scss", ".sass", ".java", ".cs", ".cpp", ".c", ".h", ".hpp", ".go", ".rs", ".php", ".rb", ".swift", ".kt", ".kts", ".sql", ".sh", ".bash", ".zsh", ".ps1", ".bat", ".dockerfile", ".xml",
    ".csv", ".tsv",
}

IMAGE_ATTACHMENT_EXTENSIONS = {
    ".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp", ".tiff", ".tif",
}

EXCEL_ATTACHMENT_EXTENSIONS = {".xlsx", ".xlsm"}
PDF_ATTACHMENT_EXTENSIONS = {".pdf"}
DOCX_ATTACHMENT_EXTENSIONS = {".docx"}


def attachment_to_embedded_context_items(
    client: Any,
    file_path: str | Path,
    model: str,
    vision_model: str | None = None,
    vision_prompt: str | None = None,
    chunk_size: int = 1_200,
    chunk_overlap: int = 180,
) -> list[dict[str, Any]]:
    """
    Read a supported attachment, extract text/metadata, embed the extracted data,
    and return standard context items.

    This function does not mutate app state. It is intended for future /attach flows,
    skill context loading, and local retrieval.
    """
    path = Path(file_path).expanduser()

    if not path.exists():
        raise FileNotFoundError(f"Attachment not found: {path}")

    if not path.is_file():
        raise ValueError(f"Attachment is not a file: {path}")

    context_items = build_attachment_context_items(
        file_path=path,
        client=client,
        vision_model=vision_model,
        vision_prompt=vision_prompt,
        chunk_size=chunk_size,
        chunk_overlap=chunk_overlap,
    )

    if path.suffix.lower() in IMAGE_ATTACHMENT_EXTENSIONS:
        return context_items

    return vectorise_context_items(client, context_items, model)


def attachments_to_embedded_context_items(
    client: Any,
    file_paths: Sequence[str | Path],
    model: str,
    vision_model: str | None = None,
    vision_prompt: str | None = None,
    chunk_size: int = 1_200,
    chunk_overlap: int = 180,
) -> list[dict[str, Any]]:
    """
    Read and embed multiple attachments.
    """
    embedded_items: list[dict[str, Any]] = []

    for file_path in file_paths:
        embedded_items.extend(
            attachment_to_embedded_context_items(
                client=client,
                file_path=file_path,
                model=model,
                vision_model=vision_model,
                vision_prompt=vision_prompt,
                chunk_size=chunk_size,
                chunk_overlap=chunk_overlap,
            )
        )

    return embedded_items

def build_attachment_context_items(
    file_path: str | Path,
    client: Any | None = None,
    vision_model: str | None = None,
    vision_prompt: str | None = None,
    chunk_size: int = 1_200,
    chunk_overlap: int = 180,
) -> list[dict[str, Any]]:
    """
    Build chunked context items for an attachment using file-type-aware chunking.
    """
    path = Path(file_path).expanduser()
    suffix = path.suffix.lower()

    if suffix in IMAGE_ATTACHMENT_EXTENSIONS:
        return build_image_context_items(
            path=path,
            client=client,
            vision_model=vision_model,
            vision_prompt=vision_prompt,
        )

    extracted_text = read_attachment_to_string(
        file_path=path,
        client=client,
        vision_model=vision_model,
        vision_prompt=vision_prompt,
    )

    return build_context_items_from_text(
        text=extracted_text,
        source_type="attachment",
        source_path=str(path.resolve()),
        title=path.name,
        metadata=build_attachment_metadata(path),
        chunk_size=chunk_size,
        chunk_overlap=chunk_overlap,
    )

def build_image_context_items(
    path: Path,
    client: Any | None = None,
    vision_model: str | None = None,
    vision_prompt: str | None = None,
) -> list[dict[str, Any]]:
    """
    Build one full-context item for an image attachment.

    Image context is intentionally not chunked because partial visual descriptions can
    remove important relationships between objects, text, UI elements, and layout.
    """
    content = read_image_file_to_string(
        path,
        client=client,
        vision_model=vision_model,
        vision_prompt=vision_prompt,
    )
    metadata = build_attachment_metadata(path)

    return [
        {
            "source_type": "image_attachment",
            "source_path": str(path.resolve()),
            "title": path.name,
            "chunk_index": 0,
            "content": content.strip(),
            "embedding": [],
            **metadata,
        }
    ]

def build_csv_context_items(
    path: Path,
    delimiter: str,
    chunk_size: int = 1_200,
    max_rows: int = 2_000,
) -> list[dict[str, Any]]:
    """
    Build table-aware context chunks for CSV/TSV files.

    Each chunk repeats the header row so retrieved table rows keep their column meaning.
    """
    raw_text = read_text_file_to_string(path)
    rows = list(csv.reader(raw_text.splitlines(), delimiter=delimiter))

    if not rows:
        return []

    header = [str(cell).strip() for cell in rows[0]]
    data_rows = rows[1:max_rows]

    return build_table_context_items(
        path=path,
        table_name=path.name,
        header=header,
        rows=data_rows,
        metadata={**build_attachment_metadata(path), "delimiter": delimiter},
        chunk_size=chunk_size,
    )


def build_excel_context_items(
    path: Path,
    chunk_size: int = 1_200,
    max_rows_per_sheet: int = 1_000,
) -> list[dict[str, Any]]:
    """
    Build table-aware context chunks for Excel files.

    Each sheet is chunked separately and repeats its first non-empty row as header/context.
    """
    if openpyxl is None:
        raise ImportError(
            "openpyxl is required to read Excel attachments. Install it with: pip install openpyxl"
        )

    workbook = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    context_items: list[dict[str, Any]] = []

    try:
        for worksheet in workbook.worksheets:
            rows: list[list[str]] = []

            for row in worksheet.iter_rows(values_only=True):
                values = ["" if value is None else str(value).strip() for value in row]
                if any(values):
                    rows.append(values)

                if len(rows) >= max_rows_per_sheet:
                    break

            if not rows:
                continue

            header = rows[0]
            data_rows = rows[1:]

            context_items.extend(
                build_table_context_items(
                    path=path,
                    table_name=worksheet.title,
                    header=header,
                    rows=data_rows,
                    metadata={**build_attachment_metadata(path), "sheet_name": worksheet.title},
                    chunk_size=chunk_size,
                )
            )
    finally:
        workbook.close()

    return context_items


def build_table_context_items(
    path: Path,
    table_name: str,
    header: Sequence[str],
    rows: Sequence[Sequence[Any]],
    metadata: dict[str, Any],
    chunk_size: int = 1_200,
) -> list[dict[str, Any]]:
    """
    Build chunks from tabular data without splitting rows.

    Every chunk includes file/table metadata and the header row.
    """
    header_line = format_table_row(header)
    context_items: list[dict[str, Any]] = []
    current_rows: list[str] = []
    current_length = 0
    row_start = 1

    prefix = f"File: {path.name}\nTable: {table_name}\nHeader: {header_line}\n"
    prefix_length = len(prefix)

    for row_number, row in enumerate(rows, start=1):
        row_line = format_table_row(row)
        row_length = len(row_line) + 1

        if current_rows and prefix_length + current_length + row_length > chunk_size:
            context_items.append(
                build_table_context_item(
                    path=path,
                    table_name=table_name,
                    prefix=prefix,
                    rows=current_rows,
                    row_start=row_start,
                    row_end=row_number - 1,
                    chunk_index=len(context_items),
                    metadata=metadata,
                )
            )

            current_rows = []
            current_length = 0
            row_start = row_number

        if not current_rows:
            row_start = row_number

        current_rows.append(row_line)
        current_length += row_length

    if current_rows:
        context_items.append(
            build_table_context_item(
                path=path,
                table_name=table_name,
                prefix=prefix,
                rows=current_rows,
                row_start=row_start,
                row_end=row_start + len(current_rows) - 1,
                chunk_index=len(context_items),
                metadata=metadata,
            )
        )

    if not context_items:
        context_items.append(
            build_table_context_item(
                path=path,
                table_name=table_name,
                prefix=prefix,
                rows=[],
                row_start=0,
                row_end=0,
                chunk_index=0,
                metadata=metadata,
            )
        )

    return context_items


def build_table_context_item(
    path: Path,
    table_name: str,
    prefix: str,
    rows: Sequence[str],
    row_start: int,
    row_end: int,
    chunk_index: int,
    metadata: dict[str, Any],
) -> dict[str, Any]:
    row_range = f"Rows: {row_start}-{row_end}" if row_start and row_end else "Rows: header only"
    content = f"{prefix}{row_range}\n" + "\n".join(rows)

    return {
        "source_type": "attachment",
        "source_path": str(path.resolve()),
        "title": path.name,
        "chunk_index": chunk_index,
        "content": content.strip(),
        "embedding": [],
        "table_name": table_name,
        "row_start": row_start,
        "row_end": row_end,
        **metadata,
    }


def format_table_row(row: Sequence[Any]) -> str:
    return " | ".join("" if value is None else str(value).strip() for value in row)

def read_attachment_to_string(
    file_path: str | Path,
    client: Any | None = None,
    vision_model: str | None = None,
    vision_prompt: str | None = None,
) -> str:
    """
    Extract readable text from supported attachment types.
    """
    path = Path(file_path).expanduser()
    suffix = path.suffix.lower()
    name = path.name.lower()

    if name in {".env", ".gitignore"} or suffix in TEXT_ATTACHMENT_EXTENSIONS:
        if suffix == ".csv":
            return read_csv_like_file_to_string(path, delimiter=",")
        if suffix == ".tsv":
            return read_csv_like_file_to_string(path, delimiter="\t")
        if suffix == ".json":
            return read_json_file_to_string(path)
        return read_text_file_to_string(path)

    if suffix in PDF_ATTACHMENT_EXTENSIONS:
        return read_pdf_file_to_string(path)

    if suffix in EXCEL_ATTACHMENT_EXTENSIONS:
        return read_excel_file_to_string(path)

    if suffix in DOCX_ATTACHMENT_EXTENSIONS:
        return read_docx_file_to_string(path)

    if suffix in IMAGE_ATTACHMENT_EXTENSIONS:
        return read_image_file_to_string(
            path,
            client=client,
            vision_model=vision_model,
            vision_prompt=vision_prompt,
        )

    raise ValueError(f"Unsupported attachment type: {path.name}")


def build_attachment_metadata(path: Path) -> dict[str, Any]:
    mime_type, _ = mimetypes.guess_type(str(path))
    return {
        "filename": path.name,
        "extension": path.suffix.lower(),
        "mime_type": mime_type or "application/octet-stream",
        "file_size_bytes": path.stat().st_size,
    }


def read_text_file_to_string(path: Path, max_bytes: int = 5_000_000) -> str:
    if path.stat().st_size > max_bytes:
        raise ValueError(f"Text file is too large to read safely: {path.name}")

    for encoding in ("utf-8", "utf-8-sig", "latin-1"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue

    return path.read_bytes().decode("utf-8", errors="replace")


def read_json_file_to_string(path: Path) -> str:
    raw_text = read_text_file_to_string(path)
    try:
        parsed = json.loads(raw_text)
    except json.JSONDecodeError:
        return raw_text
    return json.dumps(parsed, indent=2, ensure_ascii=False)


def read_csv_like_file_to_string(path: Path, delimiter: str, max_rows: int = 5000) -> str:
    raw_text = read_text_file_to_string(path)
    lines = raw_text.splitlines()

    if not lines:
        return ""

    output: list[str] = [f"File: {path.name}"]
    reader = csv.reader(lines, delimiter=delimiter)

    for row_index, row in enumerate(reader):
        if row_index >= max_rows:
            output.append(f"... truncated after {max_rows} rows")
            break
        output.append(" | ".join(str(cell).strip() for cell in row))

    return "\n".join(output)


def read_pdf_file_to_string(path: Path) -> str:
    if pypdf is None:
        raise ImportError("pypdf is required to read PDF attachments. Install it with: pip install pypdf")

    output: list[str] = [f"File: {path.name}"]
    reader = pypdf.PdfReader(str(path))

    for page_index, page in enumerate(reader.pages, start=1):
        try:
            page_text = page.extract_text() or ""
        except Exception as error:
            page_text = f"[Could not extract page text: {error}]"

        page_text = page_text.strip()
        if page_text:
            output.append(f"\n--- Page {page_index} ---")
            output.append(page_text)

    return "\n".join(output).strip()


def read_excel_file_to_string(path: Path, max_rows_per_sheet: int = 5000) -> str:
    if openpyxl is None:
        raise ImportError("openpyxl is required to read Excel attachments. Install it with: pip install openpyxl")

    workbook = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    output: list[str] = [f"File: {path.name}"]

    try:
        for worksheet in workbook.worksheets:
            output.append(f"\n--- Sheet: {worksheet.title} ---")

            for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                if row_index > max_rows_per_sheet:
                    output.append(f"... truncated after {max_rows_per_sheet} rows")
                    break

                values = ["" if value is None else str(value) for value in row]
                if any(value.strip() for value in values):
                    output.append(" | ".join(values))
    finally:
        workbook.close()

    return "\n".join(output).strip()


def read_docx_file_to_string(path: Path) -> str:
    output: list[str] = [f"File: {path.name}"]

    with zipfile.ZipFile(path) as docx_zip:
        document_xml = docx_zip.read("word/document.xml")

    root = ET.fromstring(document_xml)
    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}

    for paragraph in root.findall(".//w:p", namespace):
        texts = [node.text or "" for node in paragraph.findall(".//w:t", namespace)]
        paragraph_text = "".join(texts).strip()
        if paragraph_text:
            output.append(paragraph_text)

    return "\n".join(output).strip()

def prompt_vision_model_direct(
    client: Any,
    prompt: str,
    image_paths: Sequence[str],
    model: str,
    temperature: float = 0.1,
    top_p: float = 0.8,
) -> str:
    """
    Prompt an Ollama vision model directly using /api/chat.

    This avoids routing image prompts through generic chat/embed helpers that may not
    support Ollama's images payload correctly.
    """
    host = str(getattr(client, "host", "") or "http://localhost:11434").rstrip("/")
    url = f"{host}/api/chat"

    encoded_images: list[str] = []

    for image_path in image_paths:
        with open(image_path, "rb") as image_file:
            encoded_images.append(
                base64.b64encode(image_file.read()).decode("utf-8")
            )

    payload = {
        "model": model,
        "stream": False,
        "messages": [
            {
                "role": "user",
                "content": prompt,
                "images": encoded_images,
            }
        ],
        "options": {
            "temperature": temperature,
            "top_p": top_p,
        },
    }

    request = urllib.request.Request(
        url,
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )

    try:
        timeout = int(getattr(client, "timeout", 180) or 180)

        with urllib.request.urlopen(request, timeout=timeout) as response:
            response_data = json.loads(response.read().decode("utf-8"))

    except urllib.error.URLError as error:
        raise ConnectionError(
            f"Could not connect to Ollama vision endpoint at {url}. "
            "Check OLLAMA_HOST/config host and that Ollama is reachable."
        ) from error

    message = response_data.get("message") or {}
    content = message.get("content") or ""

    return str(content).strip()

def read_image_file_to_string(
    path: Path,
    client: Any | None = None,
    vision_model: str | None = None,
    vision_prompt: str | None = None,
) -> str:
    metadata = read_image_metadata_to_string(path)

    if client is None or not vision_model:
        return metadata

    prompt = vision_prompt or (
        "Describe this image in high detail for retrieval. Include visible text, objects, people, "
        "UI elements, charts, tables, colours, layout, and any information useful for search."
    )
    description = prompt_vision_model_direct(
        client=client,
        prompt=prompt,
        image_paths=[str(path)],
        model=vision_model,
        temperature=0.1,
        top_p=0.8,
    )

    if not description:
        return metadata

    return f"{metadata}\n\nVision description:\n{description}".strip()


def read_image_metadata_to_string(path: Path) -> str:
    lines = [f"File: {path.name}", f"Image path: {path.resolve()}"]

    if Image is None:
        lines.append("Image metadata unavailable because Pillow is not installed.")
        return "\n".join(lines)

    try:
        with Image.open(path) as image:
            lines.append(f"Format: {image.format}")
            lines.append(f"Mode: {image.mode}")
            lines.append(f"Size: {image.width}x{image.height}")
    except Exception as error:
        lines.append(f"Could not read image metadata: {error}")

    return "\n".join(lines)


def scrape_url_to_embedded_context_items(
    client: Any,
    url: str,
    model: str,
    title: str | None = None,
    timeout_ms: int = 30_000,
    chunk_size: int = 1_200,
    chunk_overlap: int = 180,
) -> list[dict[str, Any]]:
    """
    Scrape one webpage and return standard context items with embeddings applied.
    """
    page_text = scrape_url_to_string(url, timeout_ms=timeout_ms)
    context_items = build_context_items_from_text(
        text=page_text,
        source_type="webpage",
        source_path=url,
        title=title or url,
        metadata={"url": url},
        chunk_size=chunk_size,
        chunk_overlap=chunk_overlap,
    )
    return vectorise_context_items(client, context_items, model)

def duckduckgo_search_results(
    query: str,
    max_results: int = 10,
    timeout_ms: int = 30_000,
) -> list[dict[str, str]]:
    """
    Search DuckDuckGo and return normalised result records without scraping pages.
    """
    return _duckduckgo_search_results(
        query=query,
        max_results=max_results,
        timeout_ms=timeout_ms,
    )


def duckduckgo_search_results_to_embedded_context_items(
    client: Any,
    query: str,
    model: str,
    search_results: Sequence[dict[str, str]],
    timeout_ms: int = 30_000,
    chunk_size: int = 1_200,
    chunk_overlap: int = 180,
) -> list[dict[str, Any]]:
    """
    Scrape selected DuckDuckGo result records and return embedded context items.
    """
    embedded_items: list[dict[str, Any]] = []

    for result_index, result in enumerate(search_results, start=1):
        title = result.get("title") or "Untitled result"
        url = result.get("url") or ""
        snippet = result.get("snippet") or ""

        if not url:
            continue

        try:
            page_text = scrape_url_to_string(url, timeout_ms=timeout_ms)
        except Exception as error:
            page_text = f"Snippet: {snippet}\nScrape failed: {error}".strip()

        source_text = f"Title: {title}\nURL: {url}\nSnippet: {snippet}\n\n{page_text}".strip()
        context_items = build_context_items_from_text(
            text=source_text,
            source_type="search_result",
            source_path=url,
            title=title,
            metadata={
                "url": url,
                "search_query": query,
                "search_result_index": result_index,
                "snippet": snippet,
            },
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
        )
        embedded_items.extend(vectorise_context_items(client, context_items, model))

    return embedded_items

def duckduckgo_search_to_embedded_context_items(
    client: Any,
    query: str,
    model: str,
    max_results: int = 10,
    timeout_ms: int = 30_000,
    chunk_size: int = 1_200,
    chunk_overlap: int = 180,
) -> list[dict[str, Any]]:
    """
    Search DuckDuckGo, scrape result pages, and return embedded context items.

    This is intentionally structured data, unlike duckduckgo_search_and_scrape which
    returns one printable string. Main.py can add these items to VectorMemoryIndex,
    then search them against the user's prompt embedding.
    """
    search_results = _duckduckgo_search_results(
        query=query,
        max_results=max_results,
        timeout_ms=timeout_ms,
    )

    if not search_results:
        return []

    return duckduckgo_search_results_to_embedded_context_items(
        client=client,
        query=query,
        model=model,
        search_results=search_results,
        timeout_ms=timeout_ms,
        chunk_size=chunk_size,
        chunk_overlap=chunk_overlap,
    )


# -----------------------------
# Web loading utilities used by embedded context builders
# -----------------------------

COMMON_HEADERS: dict[str, str] = {
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "en-SG,en;q=0.9,en-US;q=0.8",
    "Cache-Control": "no-cache",
    "Pragma": "no-cache",
    "Upgrade-Insecure-Requests": "1",
}

COMMON_BROWSER_SETTINGS: dict[str, Any] = {
    "headless": True,
}

COMMON_CONTEXT_SETTINGS: dict[str, Any] = {
    "user_agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "viewport": {"width": 1440, "height": 900},
    "locale": "en-SG",
    "timezone_id": "Asia/Singapore",
    "java_script_enabled": True,
    "extra_http_headers": COMMON_HEADERS,
}

COMMON_PAGE_SETTINGS: dict[str, Any] = {
    "default_timeout_ms": 30_000,
    "default_navigation_timeout_ms": 30_000,
}

DUCKDUCKGO_SEARCH_URL = "https://duckduckgo.com/html/?{query}"


def create_common_browser(playwright: Any) -> Browser:
    """
    Create a Chromium browser using common browser settings.
    """
    return playwright.chromium.launch(**COMMON_BROWSER_SETTINGS)


def create_common_context(browser: Browser, **overrides: Any) -> BrowserContext:
    """
    Create a browser context using common headers, user agent, locale, timezone, viewport,
    and JavaScript settings.

    Args:
        browser: A Playwright browser instance.
        **overrides: Optional context settings to override the defaults.

    Returns:
        A configured Playwright browser context.
    """
    settings = {**COMMON_CONTEXT_SETTINGS, **overrides}
    return browser.new_context(**settings)


def create_common_page(context: BrowserContext, timeout_ms: int | None = None) -> Page:
    """
    Create a page using common timeout settings.

    Args:
        context: A Playwright browser context.
        timeout_ms: Optional timeout override in milliseconds.

    Returns:
        A configured Playwright page.
    """
    page = context.new_page()
    page.set_default_timeout(timeout_ms or COMMON_PAGE_SETTINGS["default_timeout_ms"])
    page.set_default_navigation_timeout(
        timeout_ms or COMMON_PAGE_SETTINGS["default_navigation_timeout_ms"]
    )
    return page


def scrape_url_to_string(url: str, timeout_ms: int = 30_000) -> str:
    """
    Open a webpage with Playwright, scrape the visible page text, and return it as a string.

    The page is loaded with the common browser/context/page settings, including common
    headers and JavaScript enabled.

    Args:
        url: The website URL to access.
        timeout_ms: Maximum time to wait for page navigation and loading, in milliseconds.

    Returns:
        The visible text content of the webpage as a cleaned string.
    """
    if not url or not url.strip():
        raise ValueError("url must be a non-empty string")

    with sync_playwright() as playwright:
        browser = create_common_browser(playwright)
        context = create_common_context(browser)
        page = create_common_page(context, timeout_ms=timeout_ms)

        try:
            page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
            _wait_for_page_ready(page, timeout_ms=timeout_ms)
            page_text = page.locator("body").inner_text(timeout=timeout_ms)
            return _clean_scraped_text(page_text)
        finally:
            context.close()
            browser.close()

class DuckDuckGoHTMLParser(HTMLParser):
    """
    Parse DuckDuckGo's lightweight HTML search results.
    """

    def __init__(self) -> None:
        super().__init__()
        self.results: list[dict[str, str]] = []
        self._current: dict[str, str] | None = None
        self._capture_title = False
        self._capture_snippet = False
        self._title_parts: list[str] = []
        self._snippet_parts: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attr_map = {key: value or "" for key, value in attrs}
        class_name = attr_map.get("class", "")

        if tag == "a" and "result__a" in class_name:
            self._current = {
                "title": "",
                "url": _normalise_duckduckgo_result_url(attr_map.get("href", "")),
                "snippet": "",
            }
            self._capture_title = True
            self._title_parts = []
            return

        if self._current is not None and tag in {"a", "div"} and "result__snippet" in class_name:
            self._capture_snippet = True
            self._snippet_parts = []

    def handle_data(self, data: str) -> None:
        if self._capture_title:
            self._title_parts.append(data)

        if self._capture_snippet:
            self._snippet_parts.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag == "a" and self._capture_title and self._current is not None:
            self._current["title"] = _normalise_duckduckgo_text(" ".join(self._title_parts))
            self._capture_title = False
            return

        if self._capture_snippet and tag in {"a", "div"} and self._current is not None:
            self._current["snippet"] = _normalise_duckduckgo_text(" ".join(self._snippet_parts))
            self._capture_snippet = False

            if self._current.get("title") and self._current.get("url"):
                self.results.append(self._current)

            self._current = None


def _duckduckgo_search_results(
    query: str,
    max_results: int = 10,
    timeout_ms: int = 30_000,
) -> list[dict[str, str]]:
    """
    Search DuckDuckGo using the lightweight HTML endpoint and return normalised results.

    This uses the known working access path:
    urllib request -> DuckDuckGo HTML -> HTMLParser.

    The rest of the /web pipeline still uses our own scraping, chunking, embedding,
    storage, and ranking logic.
    """
    if not query or not query.strip():
        raise ValueError("query must be a non-empty string")

    encoded_query = urllib.parse.urlencode({"q": query})
    search_url = DUCKDUCKGO_SEARCH_URL.format(query=encoded_query)
    timeout_seconds = max(1, int(timeout_ms / 1000))

    request = urllib.request.Request(
        search_url,
        headers={
            "User-Agent": "Mozilla/5.0 SimpleAgent/1.0",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9",
        },
    )

    try:
        with urllib.request.urlopen(request, timeout=timeout_seconds) as response:
            html_text = response.read().decode("utf-8", errors="replace")
    except Exception:
        return []

    parser = DuckDuckGoHTMLParser()

    try:
        parser.feed(html_text)
    except Exception:
        return []

    deduped: list[dict[str, str]] = []
    seen_urls: set[str] = set()

    for result in parser.results:
        title = result.get("title", "")
        url = result.get("url", "")

        if not _is_valid_duckduckgo_result(title=title, url=url, seen_urls=seen_urls):
            continue

        seen_urls.add(url)
        deduped.append(result)

        if len(deduped) >= max_results:
            break

    return deduped

def _extract_duckduckgo_results_from_page(
    page: Page,
    max_results: int,
    seen_urls: set[str],
    timeout_ms: int,
) -> list[dict[str, str]]:
    results: list[dict[str, str]] = []

    result_links = page.locator("a.result__a").all()

    for link in result_links:
        if len(results) >= max_results:
            break

        try:
            title = link.inner_text(timeout=timeout_ms).strip()
            raw_url = link.get_attribute("href") or ""
        except Exception:
            continue

        url = _normalise_duckduckgo_result_url(raw_url)

        if not _is_valid_duckduckgo_result(title=title, url=url, seen_urls=seen_urls):
            continue

        snippet = ""
        try:
            result_container = link.locator("xpath=ancestor::*[contains(@class, 'result')]").first
            snippet_element = result_container.locator(".result__snippet").first
            if snippet_element.count() > 0:
                snippet = snippet_element.inner_text(timeout=timeout_ms).strip()
        except Exception:
            snippet = ""

        seen_urls.add(url)
        results.append(
            {
                "title": title,
                "url": url,
                "snippet": snippet,
            }
        )

    return results


def _is_valid_duckduckgo_result(title: str, url: str, seen_urls: set[str]) -> bool:
    if not title or not url or url in seen_urls:
        return False

    parsed = urlparse(url if "://" in url else f"https://{url}")
    host = parsed.netloc.lower()

    if not host:
        return False

    if host.endswith(".duckduckgo.com") or host in {
        "duckduckgo.com",
        "www.duckduckgo.com",
        "lite.duckduckgo.com",
        "html.duckduckgo.com",
    }:
        return False

    return True

def _normalise_duckduckgo_result_url(raw_url: str) -> str:
    """
    Convert DuckDuckGo redirect URLs into direct target URLs when possible.
    """
    if not raw_url:
        return ""

    decoded_url = html.unescape(raw_url.strip())

    if "uddg=" in decoded_url:
        parsed_url = urllib.parse.urlparse(decoded_url)
        query_params = urllib.parse.parse_qs(parsed_url.query)

        if "uddg" in query_params and query_params["uddg"]:
            return urllib.parse.unquote(query_params["uddg"][0])

    return decoded_url

def _normalise_duckduckgo_text(value: str) -> str:
    return re.sub(r"\s+", " ", html.unescape(value)).strip()

def _wait_for_page_ready(page: Page, timeout_ms: int) -> None:
    """
    Wait for the page to become reasonably ready without failing hard when a website keeps
    background network requests open.
    """
    try:
        page.wait_for_load_state("networkidle", timeout=timeout_ms)
    except Exception:
        page.wait_for_load_state("load", timeout=timeout_ms)


def _clean_scraped_text(text: str) -> str:
    """
    Normalise scraped text by removing excessive blank lines and whitespace.
    """
    lines = [line.strip() for line in text.splitlines()]
    cleaned_lines = [line for line in lines if line]
    return "\n".join(cleaned_lines)